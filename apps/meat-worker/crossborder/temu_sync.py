"""Temu read-only sync adapter extracted from SaaS-HZ's Playwright crawler."""

from __future__ import annotations

import os
import sys
from pathlib import Path
from typing import Any

from crossborder.contract import SyncRequest


def _vendor_root() -> Path:
    return Path(__file__).resolve().parent / "vendor"


def _tenant_id(account_ref: str) -> int:
    if not account_ref.isdigit() or int(account_ref) <= 0:
        raise ValueError("Temu account_ref must be the positive local profile ID")
    return int(account_ref)


def run(request: SyncRequest, cfg: Any) -> dict[str, Any]:
    try:
        tenant_id = _tenant_id(request.account_ref)
    except ValueError as exc:
        return {"ok": False, "need_login": False, "error": str(exc)}
    vendor_root = _vendor_root()
    if str(vendor_root) not in sys.path:
        sys.path.insert(0, str(vendor_root))
    os.environ["TEMU_PROFILE_ROOT"] = (
        os.environ.get("CROSSBORDER_TEMU_PROFILE_ROOT", "").strip()
        or str(cfg.profile_dir() / "crossborder" / "temu")
    )
    os.environ["TEMU_HEADLESS"] = "1" if not cfg.headed else "0"
    os.environ.setdefault("TEMU_BROWSER_CHANNEL", "chrome" if cfg.use_system_chrome else "")
    if request.scope == "activity_data":
        return _export_activity_data(tenant_id, cfg)
    try:
        from app.crawler.temu_crawler import crawl_temu_sales_live

        payload = crawl_temu_sales_live(request.date_end or None, tenant_id=tenant_id)
    except Exception as exc:  # noqa: BLE001
        message = str(exc)
        return {
            "ok": False,
            "need_login": "登录" in message or "login" in message.lower(),
            "error": message[:500],
        }
    rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    shops = payload.get("shops") if isinstance(payload.get("shops"), list) else []
    return {
        "ok": True,
        "summary": {
            "report_time": payload.get("report_time"),
            "shop_count": len(shops),
            "row_count": len(rows),
        },
        "diagnostics": {"source": "temu_seller", "mode": "playwright"},
    }


def _export_activity_data(tenant_id: int, cfg: Any) -> dict[str, Any]:
    from openpyxl import load_workbook
    from app.browser.context import get_or_open_seller_page, open_temu_context

    export_dir = cfg.profile_dir() / "crossborder" / "exports" / "temu" / str(tenant_id)
    export_dir.mkdir(parents=True, exist_ok=True)
    activity_url = "https://agentseller.temu.com/main/act/data-full"
    try:
        with open_temu_context(tenant_id, headless=not cfg.headed) as (_, context):
            page = get_or_open_seller_page(context)
            page.goto(activity_url, wait_until="domcontentloaded", timeout=120_000)
            page.wait_for_timeout(3_000)
            page.evaluate(
                '''() => { const label="\\u6279\\u91cf\\u5bfc\\u51fa\\u6570\\u636e"; const el=Array.from(document.querySelectorAll('button,a,[role="button"],span')).find(e=>(e.innerText||'').trim()===label); if(!el) throw new Error('Temu activity export button unavailable'); el.click(); }'''
            )
            page.wait_for_timeout(500)
            boxes = page.locator('label[data-testid="beast-core-checkbox"]')
            if boxes.count() < 2:
                return {"ok": False, "error": "Temu activity export has no selectable detail rows"}
            boxes.nth(0).click()
            page.wait_for_timeout(250)
            with page.expect_download(timeout=120_000) as download_info:
                page.evaluate(
                    '''() => { const label="\\u786e\\u8ba4"; const buttons=Array.from(document.querySelectorAll('button')).filter(e=>(e.innerText||'').trim()===label); const button=buttons.at(-1); if(!button) throw new Error('Temu activity export confirmation unavailable'); button.click(); }'''
                )
            download = download_info.value
            file_path = export_dir / download.suggested_filename
            download.save_as(str(file_path))
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": str(exc)[:500]}
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value or "") for value in rows[0]] if rows else []
    return {
        "ok": True,
        "summary": {"file_name": file_path.name, "row_count": max(0, len(rows) - 1), "columns": headers},
        "diagnostics": {"source": "temu_activity_export", "format": file_path.suffix.lower()},
    }
